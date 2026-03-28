"""Export builders for API download endpoints."""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from collections import Counter
from typing import Any, Dict, List, Literal, Tuple

from modules.review_db import ReviewDB

ExportFormat = Literal["json", "csv", "xlsx"]

CSV_COLUMNS: List[str] = [
    "place_id",
    "place_name",
    "review_id",
    "author",
    "rating",
    "review_text_primary",
    "review_text_all_json",
    "review_date",
    "raw_date",
    "likes",
    "profile_url",
    "is_deleted",
    "created_date",
    "last_modified",
    "last_seen_session",
    "last_changed_session",
    "owner_responses_json",
    "user_images_json",
    "s3_images_json",
    "source_url",
    "resolved_place_url",
    "scrape_session_id",
    "scrape_started_at",
    "scrape_completed_at",
    "scrape_mode",
    "google_maps_auth_mode",
    "sort_order_requested",
    "sort_order_confirmed",
    "extraction_confidence",
    "source_locale",
]


INVALID_SHEET_CHARS = re.compile(r'[:\\/*?\[\]]')


def _safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "")).strip("._")
    return cleaned or "place"


def _safe_sheet_name(base_name: str, existing: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", str(base_name or "")).strip() or "sheet"
    cleaned = cleaned[:31]
    if cleaned not in existing:
        existing.add(cleaned)
        return cleaned

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in existing:
            existing.add(candidate)
            return candidate
        i += 1


def _to_json_bytes(payload: Dict[str, Any]) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _filter_empty_text(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Remove rows where review_text_primary is empty/None (star-only reviews)."""
    return [
        row for row in rows
        if (row.get("review_text_primary") or "").strip()
    ]


def _filter_min_review_count_per_place(
    payload: Dict[str, Any],
    rows: List[Dict[str, Any]],
    min_review_count: int | None,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    threshold = max(0, int(min_review_count or 0))
    if threshold <= 0:
        return payload, rows

    counts = Counter(str(row.get("place_id") or "") for row in rows if str(row.get("place_id") or ""))
    keep_place_ids = {place_id for place_id, count in counts.items() if count >= threshold}

    filtered_rows = [
        row for row in rows
        if str(row.get("place_id") or "") in keep_place_ids
    ]

    next_payload = dict(payload)
    next_payload["places"] = [
        place for place in list(payload.get("places") or [])
        if str(place.get("place_id") or "") in keep_place_ids
    ]
    next_payload["reviews_by_place"] = {
        place_id: place_rows
        for place_id, place_rows in dict(payload.get("reviews_by_place") or {}).items()
        if str(place_id or "") in keep_place_ids
    }
    export_meta = dict(next_payload.get("export_meta") or {})
    export_meta["min_review_count"] = threshold
    next_payload["export_meta"] = export_meta
    return next_payload, filtered_rows


def _to_csv_bytes(rows: List[Dict[str, Any]], columns: List[str] | None = None) -> bytes:
    cols = columns or CSV_COLUMNS
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    # Emit UTF-8 BOM so spreadsheet apps recognize Chinese and other non-ASCII
    # text correctly when users open downloaded CSV files directly.
    return buf.getvalue().encode("utf-8-sig")


def _to_xlsx_single_place(payload: Dict[str, Any], rows: List[Dict[str, Any]], sheet_name: str | None = None, columns: List[str] | None = None) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required for XLSX export") from e

    place = payload.get("place", {}) if isinstance(payload, dict) else {}
    meta = payload.get("export_meta", {}) if isinstance(payload, dict) else {}

    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(["field", "value"])
    summary.append(["place_id", place.get("place_id", "")])
    summary.append(["place_name", place.get("place_name", "")])
    summary.append(["original_url", place.get("original_url", "")])
    summary.append(["resolved_url", place.get("resolved_url", "")])
    summary.append(["last_scraped", place.get("last_scraped", "")])
    summary.append(["generated_at", meta.get("generated_at", "")])
    summary.append(["scope", meta.get("scope", "")])
    summary.append(["include_deleted", str(meta.get("include_deleted", False))])
    summary.append(["db_path_basename", meta.get("db_path_basename", "")])
    summary.append(["active_conflict_groups", meta.get("active_conflict_groups", 0)])

    cols = columns or CSV_COLUMNS
    reviews_sheet_name = sheet_name or "reviews"
    reviews_sheet_name = INVALID_SHEET_CHARS.sub("_", reviews_sheet_name).strip()[:31] or "reviews"
    ws = wb.create_sheet(reviews_sheet_name)
    ws.append(cols)
    for row in rows:
        ws.append([row.get(name, "") for name in cols])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _to_xlsx_all_places(payload: Dict[str, Any], rows: List[Dict[str, Any]], sheet_name: str | None = None, columns: List[str] | None = None) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required for XLSX export") from e

    places = payload.get("places", []) if isinstance(payload, dict) else []

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("place_id") or ""), []).append(row)

    wb = Workbook()
    index = wb.active
    index_title = sheet_name or "index"
    index_title = INVALID_SHEET_CHARS.sub("_", index_title).strip()[:31] or "index"
    index.title = index_title
    index.append(["place_id", "place_name", "row_count"])

    cols = columns or CSV_COLUMNS
    used_sheet_names = {index_title}
    for place in places:
        place_id = str(place.get("place_id") or "")
        place_name = str(place.get("place_name") or "")
        place_rows = grouped.get(place_id, [])
        index.append([place_id, place_name, len(place_rows)])

        per_place_sheet_base = place_name or place_id or "place"
        per_place_sheet_name = _safe_sheet_name(per_place_sheet_base, used_sheet_names)
        ws = wb.create_sheet(per_place_sheet_name)
        ws.append(cols)
        for row in place_rows:
            ws.append([row.get(name, "") for name in cols])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _validate_columns(columns: List[str] | None) -> List[str] | None:
    """Return only columns that exist in CSV_COLUMNS, preserving order."""
    if not columns:
        return None
    valid = [c for c in columns if c in CSV_COLUMNS]
    return valid or None


def build_place_export(
    review_db: ReviewDB,
    place_id: str,
    fmt: ExportFormat,
    include_deleted: bool = False,
    exclude_empty_text: bool = False,
    sheet_name: str | None = None,
    columns: List[str] | None = None,
) -> Tuple[bytes, str, str]:
    payload = review_db.export_place_json_payload(place_id, include_deleted=include_deleted)
    rows = review_db.export_place_flat_rows(place_id, include_deleted=include_deleted)
    if exclude_empty_text:
        rows = _filter_empty_text(rows)

    cols = _validate_columns(columns)
    safe_pid = _safe_name(place_id)
    if fmt == "json":
        return _to_json_bytes(payload), "application/json", f"reviews_{safe_pid}.json"
    if fmt == "csv":
        return _to_csv_bytes(rows, columns=cols), "text/csv; charset=utf-8", f"reviews_{safe_pid}.csv"
    if fmt == "xlsx":
        return (
            _to_xlsx_single_place(payload, rows, sheet_name=sheet_name, columns=cols),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"reviews_{safe_pid}.xlsx",
        )
    raise ValueError(f"Unsupported export format: {fmt}")


def build_all_export(
    review_db: ReviewDB,
    fmt: ExportFormat,
    include_deleted: bool = False,
    exclude_empty_text: bool = False,
    min_review_count: int | None = None,
    sheet_name: str | None = None,
    columns: List[str] | None = None,
) -> Tuple[bytes, str, str]:
    payload = review_db.export_all_json_payload(include_deleted=include_deleted)
    rows = review_db.export_all_flat_rows(include_deleted=include_deleted)
    if exclude_empty_text:
        rows = _filter_empty_text(rows)
    payload, rows = _filter_min_review_count_per_place(payload, rows, min_review_count)

    cols = _validate_columns(columns)
    if fmt == "json":
        return _to_json_bytes(payload), "application/json", "reviews_all.json"
    if fmt == "csv":
        return _to_csv_bytes(rows, columns=cols), "text/csv; charset=utf-8", "reviews_all.csv"
    if fmt == "xlsx":
        return (
            _to_xlsx_all_places(payload, rows, sheet_name=sheet_name, columns=cols),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "reviews_all.xlsx",
        )
    raise ValueError(f"Unsupported export format: {fmt}")


def db_path_basename(review_db: ReviewDB) -> str:
    return Path(str(review_db.backend.db_path)).name
