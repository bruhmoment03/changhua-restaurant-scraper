"""Tests for tools/csv_exports_to_xlsx.py."""

from __future__ import annotations

import csv
import importlib.util
import sys
from pathlib import Path

import pytest


def _load_tool_module(module_name: str, rel_path: str):
    root = Path(__file__).resolve().parents[1]
    path = root / rel_path
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["review_id", "author", "rating"])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def test_convert_csv_dir_to_xlsx_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    mod = _load_tool_module("csv_exports_to_xlsx_test_mod", "tools/csv_exports_to_xlsx.py")

    csv_dir = tmp_path / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    _write_csv(
        csv_dir / "reviews_place_1.csv",
        [
            {"review_id": "r1", "author": "A", "rating": "5"},
            {"review_id": "r2", "author": "B", "rating": "4"},
        ],
    )
    _write_csv(
        csv_dir / "reviews_place_2.csv",
        [{"review_id": "r3", "author": "C", "rating": "3"}],
    )

    out_path = tmp_path / "exports" / "reviews.xlsx"
    total = mod.convert_csv_dir_to_xlsx(str(csv_dir), str(out_path))
    assert total == 3
    assert out_path.exists()

    wb = load_workbook(out_path)
    assert "index" in wb.sheetnames
    idx = wb["index"]
    rows = list(idx.iter_rows(values_only=True))
    assert rows[0] == ("place_id", "sheet_name", "row_count", "csv_file")
    assert len(rows) == 3
    assert sorted([rows[1][0], rows[2][0]]) == ["place_1", "place_2"]
    assert sorted([rows[1][2], rows[2][2]]) == [1, 2]


def test_safe_sheet_name_handles_collisions():
    mod = _load_tool_module("csv_exports_to_xlsx_test_mod_names", "tools/csv_exports_to_xlsx.py")
    existing = {"index"}
    base = "x" * 40
    first = mod._safe_sheet_name(base, existing)
    second = mod._safe_sheet_name(base, existing)
    assert first != second
    assert len(first) <= 31
    assert len(second) <= 31


def test_convert_requires_csv_files(tmp_path):
    mod = _load_tool_module("csv_exports_to_xlsx_test_mod_empty", "tools/csv_exports_to_xlsx.py")
    csv_dir = tmp_path / "empty_csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    with pytest.raises(ValueError, match="No CSV files found"):
        mod.convert_csv_dir_to_xlsx(str(csv_dir), str(tmp_path / "out.xlsx"))
