#!/usr/bin/env python3
"""Convert per-place CSV exports into a single XLSX workbook."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path


INVALID_SHEET_CHARS = re.compile(r'[:\\/*?\[\]]')


def _extract_place_id(path: Path) -> str:
    stem = path.stem
    if stem.startswith("reviews_"):
        return stem[len("reviews_"):] or stem
    return stem


def _safe_sheet_name(base_name: str, existing: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", base_name).strip() or "sheet"
    cleaned = cleaned[:31]
    if cleaned not in existing:
        existing.add(cleaned)
        return cleaned

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in existing:
            existing.add(candidate)
            return candidate
        i += 1


def convert_csv_dir_to_xlsx(csv_dir: str, out_path: str) -> int:
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        ) from e

    input_dir = Path(csv_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"CSV directory not found: {csv_dir}")
    csv_files = sorted(input_dir.glob("*.csv"))
    if not csv_files:
        raise ValueError(f"No CSV files found in: {csv_dir}")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    index_ws = wb.create_sheet("index")
    index_ws.append(["place_id", "sheet_name", "row_count", "csv_file"])

    existing_sheets: set[str] = {"index"}
    total_rows = 0

    for csv_file in csv_files:
        place_id = _extract_place_id(csv_file)
        sheet_name = _safe_sheet_name(place_id, existing_sheets)
        ws = wb.create_sheet(sheet_name)

        row_count = 0
        with csv_file.open("r", encoding="utf-8", newline="") as fh:
            reader = csv.DictReader(fh)
            if reader.fieldnames:
                ws.append(reader.fieldnames)
            for row in reader:
                ws.append([row.get(name, "") for name in (reader.fieldnames or [])])
                row_count += 1

        index_ws.append([place_id, sheet_name, row_count, str(csv_file)])
        total_rows += row_count

    wb.save(out)
    return total_rows


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description="Convert CSV exports (one file per place) into a single XLSX workbook."
    )
    parser.add_argument(
        "--csv-dir",
        default="exports",
        help="Directory containing CSV files (default: exports)",
    )
    parser.add_argument(
        "--out",
        default="exports/reviews.xlsx",
        help="Output XLSX file path (default: exports/reviews.xlsx)",
    )
    args = parser.parse_args(argv)

    rows = convert_csv_dir_to_xlsx(args.csv_dir, args.out)
    print(f"Wrote {args.out} with {rows} data row(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(__import__("sys").argv[1:]))
